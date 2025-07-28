import matplotlib.pyplot as plt
from openpyxl import load_workbook
from Data import countStep

#для вывода трасс
def createChartWayThree(name, way):
    mass = [[], [], []]
    fig, axs = plt.subplots(nrows=3, ncols=1)
    index = []

    for i in range(30):
        index.append(i)


    for i in range(3):
        point = 0
        for x in way[i]:
            point += x
            mass[i].append(point)

    for i in range(3):
        axs[i].plot(index, mass[i])

    plt.suptitle(name)
    plt.show()

#основной график со всеми результатами по всем трасса
def createChartThree(name, goodGen, avgGoodGen, factGen, avgBestThreeWay):
    fig, axs = plt.subplots(nrows=4, ncols=1)  # Создаем фигуру (необязательно, но позволяет задать размер)
    listStep = []
    listFact = [[], [], [], [], [], [], [], [], []]

    for i in range(countStep):
        listStep.append(i + 1)
        for i in range(len(factGen)):
            listFact[i].append(factGen[i])

    for i in range(3):
        axs[i].plot(listStep, goodGen[i], label='Значение лучшей особи на трассе', color='blue')
        axs[i].plot(listStep, avgGoodGen[i], label='Среднее значение 10 лучших особей на трассе', color='red')
        axs[i].plot(listStep, listFact[i], label='Фиксированная особь', color='green')
        axs[i].legend()

    axs[3].plot(listStep, avgBestThreeWay, label='Лучшее среднее значение по трём трассам среди особей', color='red')
    axs[3].legend()

    plt.suptitle(name)
    plt.show()

def createTableBestGenotype(genotype, name):
    wb = load_workbook("Generation/best.xlsx")

    ws = wb.create_sheet(title=name)

    for row in genotype:
        ws.append(row)

    wb.save("Generation/best.xlsx")

