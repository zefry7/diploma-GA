from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook

#Запись в файл
def WriteInFiles(result_distance, n, rating, populationList):
    f = open(f"Generation/Generation-{n}.txt", "w")
    f.write("Результаты по дистанциям:\n")
    f.write(result_distance)
    f.write("Рейтинг после прохождения:\n")
    f.write(f"{rating}")
    f = open(f"Generation/Gen-{n}.txt", "w")

    str = ""
    for x in populationList:
        str += f"{x}\n"
    f.write(str)


#Запись в Excel
def WriteExcel(info, col_number, dop_col):
    filename = f"Generation/result.xlsx"
    flag = False
    try:
        workbook = load_workbook(filename)
        sheet = workbook.active
    except FileNotFoundError:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Результаты"

    for row_index, row_data in enumerate(info):
        for col_index, col_data in enumerate(row_data):
            if col_data == 'Контрольные трассы':
                flag = True
            if flag == False:
                sheet.cell(row=row_index + 1 + col_number, column=col_index + 1 + dop_col).value = col_data
            else:
                sheet.cell(row=row_index + 1 + col_number - 11, column=col_index + 1 + 10).value = col_data

    workbook.save(filename)

